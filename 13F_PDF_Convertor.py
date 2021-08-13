'''****************************************************************************************
13F PDF convertor tool

Purpose: convert 13F quarterly PDF file into an Excel file for further work processing.

Input: 13F PDF file from SEC website https://www.sec.gov/divisions/investment/13flists.htm

Output: 13F list in Excel format with some limitations

Limitations:
1) Some rows don't have the split between Description and Issue columns

when I used tika lib to convert PDF to raw string, the table structure was lost.
So I had to split issuer name and issue description column (combined and called "description"
in the following code explanation) using some identifiers (called "starter" and "end" words).
Since the issuer name content has some cut-off names due to limited column width, using my
identifiers cannot 100% separate issuer name and issuer description. For the rows still with
combined issuer description, there is a note "See description column for issue type" in Issue
column of the final output Excel file. Out of Q2 2021 13F list, only 278 rows out of total
21312 rows(1.3%) have this limitation.

2) Incorrect split between issuer name description and issue security name description

There were some rare situations that the issuer name (in Description column) and issue
(in Issue column) were split incorrectly. The common reason is that the issued security starts
with an uncommon word which were not included in the issue starter word list under check_starter
help function (below) or the issuer's name ends with an uncommon word which were not included in
the issuer end word list under check_end help function (below). But the occurrences represent less
than 1% of the total population.

This output Excel file is not perfect, but it is good enough for my work process. If you have
any improvement idea, please let me know on GitHub. Happy to discuss and make it perfect.

*******************************************************************************************'''

import tkinter as tk
from tkinter import filedialog
from os import getcwd
from glob import glob
from tika import parser
import re
import pandas as pd
import openpyxl

'''
Use a tk window to obtain PDF file path
Input: none
Output: file path in string like this (on mac) "/Users/xxxx/Documents/Project/13F/13F_PDF_files/13flist2021q2.pdf
'''


def obtain_pdf_file_path():
    root = tk.Tk()
    root.withdraw()  # hides the root window

    # create window to select PDF file
    root.filename =  filedialog.askopenfilename(initialdir=getcwd(),\
                                                title = "Select file",\
                                                filetypes = (("PDF file","*.pdf"),\
                                                             ("all files","*.*")))
    return root.filename


def save_excel_file_path():
    root = tk.Tk()
    root.withdraw()  # hides the root window

    # create window to select PDF file
    root.filename = filedialog.asksaveasfilename(initialdir=getcwd(), \
                                                 title="Select file", \
                                                 filetypes=(("Excel files", "*.xlsx"), \
                                                            ("all files", "*.*")))
    return root.filename


'''
Read PDF into string and organize the string by line into a list
Input: file path
Output: a split list of strings by PDF line without empty lines
'''


def parsePDF(input_path):
    for input_file in glob(input_path):
        # make file into a string
        parserPDF = parser.from_file(input_file)
        # make string more readable separating by line (it is still a str)
        pdf = parserPDF['content']
        # split the str by line break (\n) and add each block of string in a list
        split = pdf.splitlines()
        # remove empty line
        split2 = [x for x in split if x.strip()]

        return split2


'''
Categorize string lines into these categories: CUSIP, *, description, issue and status (ADDED or DELETED)
Input: one string from the split list returned from parsePDF function above
Output: a list of one row in the order of CUSIP, *, description, issue and status (from left to right)
'''


# Help function: cut CUSIP part from the string(el) and return CUSIP string
def add_CUSIP(match_CUSIP, el):

    # Find start and end indexes of CUSIP from the string(el)
    index_CUSIP = match_CUSIP.span()
    # Add CUSIP part of the string into row_list by tracing the start and end indexes
    return el[index_CUSIP[0]: index_CUSIP[1]]


# Help function: if there is "*" following the CUSIP string, return "*", otherwise, return ""
def add_star(match_CUSIP, el):

    # Find start and end indexes of CUSIP from the string(el)
    index_CUSIP = match_CUSIP.span()
    if el[index_CUSIP[1]+1] == "*":
        return "*"
    return ""

# Helper in add_description help function: Find start index of the description based on if the row has "*"

def description_start_index_check_star(index_CUSIP, el):

    if el[index_CUSIP[1] + 1] == "*":
        return index_CUSIP[1] + 3
    return index_CUSIP[1] + 1


# Help function: find description part from the string (description AND issue columns)
def add_description(match_CUSIP, el):

    # Find start and end indexes of CUSIP from the string(el)
    index_CUSIP = match_CUSIP.span()

    # Check if the string(el) contains status (ADDED or DELETED)
    match_status1 = re.search('ADDED$', el)
    match_status2 = re.search('DELETED$', el)

    # Find description string by cutting off "*" on its left and status (ADDED or DELETED) on its right side
    if match_status1 == None and match_status2 == None:
        return el[description_start_index_check_star(index_CUSIP,el):].strip()
    elif match_status1 != None:
        index_status1 = match_status1.span()
        return el[description_start_index_check_star(index_CUSIP, el):index_status1[0]].strip()
    else:
        index_status2 = match_status2.span()
        return el[description_start_index_check_star(index_CUSIP, el):index_status2[0]].strip()

# Help function: Find status string from the end of the string(el), if no status, return ""
def add_status(match_CUSIP, el):

    # Find start and end indexes of CUSIP from the string(el)
    index_CUSIP = match_CUSIP.span()

    # Check if the string(el) contains status (ADDED or DELETED)
    match_status1 = re.search('ADDED$', el)
    match_status2 = re.search('DELETED$', el)

    # Find status string from the end of the string(el), if no status, return ""
    if match_status1 == None and match_status2 == None:
        return ""
    elif match_status1 != None:
        index_status1 = match_status1.span()
        return el[index_status1[0]: index_status1[1]].strip()
    else:
        index_status2 = match_status2.span()
        return el[index_status2[0]: index_status2[1]].strip()

# Input row info (CUSIP, *, description and status) into a list
# Note: the description here includes BOTH description and issue columns. See split_description function below
# for further modification
def categorize_col_contents(el):
    row_list = []
    # check if the string (el) contains CUSIP expression in [digit....whitespace..whitespace.]
    match_CUSIP = re.match(r'.\d{1}....\s.{2}\s.', el)

    if match_CUSIP == None: # Check if the string contains useful info (with a CUSIP)
        return None
    else:
        row_list.append(add_CUSIP(match_CUSIP, el))
        row_list.append(add_star(match_CUSIP, el))
        row_list.append(add_description(match_CUSIP, el))
        row_list.append(add_status(match_CUSIP, el))
        return row_list

'''
Further split description and issue columns based on common start words of the issue column and common end words of the 
description column. 
Input: all_table: the list containing all rows in lists (CUSIP, *, description and status); the description in all_table
including strings of both description and issue columns in the actual 13F table
Output: final_table: the final list containing all rows in lists (CUSIP, *, description, issue and status)
'''

# Helper of the help function to check if the common words of issue in all_table's description string
def starter_check(des):
    # summarize common words at the beginning of issue column content (THE ORDER IN THIS LIST MATTER)
    starters = ["CALL", "PUT", "DEBT", "NAMEN", "*W", "ORDINARY", "RIGHT", "WBI",
                "NOTE", "USD ORD", "ORD SHS", "ORD SH", "USD MFC", "REG SHS", "SPONS ADR", "SPONS ADS", "SPON ADR",
                "SPON ADS", "SPONS ADS", "SPONSORD ADS", "SPONSORED", "SPONDS", "PHYSCL", "PRTNRSP", "PARTNERSHP",
                "SHS CL", "COM CL", "COM CLASS", " ORD","COM STK", "COM UNIT", "CL A", "USD ORD", "ORD SH", "SHS CLASS",
                "UNIT COM", "ADR", "ADS", "COM", "UNIT", "ORD", "SHS", "CLASS", "S&P"]

    # iterate through starters list and return a list of:
    # 1) True/False (if the des contains any starter word);
    # 2) if True for 1), return the actual starter word in string, otherwise, return None.

    for starter in starters:
        if " "+starter in des:
            return [True, starter]
    return [False, None]

# Helper of the help function to check if the common words of description in all_table's description string
# Note: all_table's description (as the input) contains both description and issue columns' content
def end_check(des):
    # summarize common words at the end of description column content (THE ORDER IN THIS LIST MATTER)
    ends = ["CORP N", "CORP NEW", "INC N", "INC NEW", "PLC", "LTD", "INC", "CORP", "HLDGS I", "HLDGS II", "HLDGS III",
            "HLDGS", "FD TR", "ETF TR", "BRH", "L P", " LP", "S A", "TR I", "TR II", "TR III", "FD I", "FD II",
            "FD III", "ETF TR", "FD T", "TRADED FD", "FD I", "FD II", "FD III", "TR", "FDS", "FD", "TRUST"]
    # iterate through ends list and return a list of:
    # 1) True/False (if the des contains any end word);
    # 2) if True for 1), return the actual end word in string, otherwise, return None.

    reverse_des = des[::-1]

    for end in ends:
        reverse_end = end[::-1]
        if " "+reverse_end+" " in reverse_des:
            return [True, end.strip()]
    return [False, None]

# Further split description and issue columns based on common start words of the issue column and common end words of
# the description column.
def split_table(all_table):

    for row in all_table:
        des = row[2] # locate all_table's description string currently stored in each list (or called row here)
        if starter_check(des)[0]:  # if the des string contain any starter word
            # locate the start index of starter in the des string
            starter_start_index = des.rfind(starter_check(des)[1])
            # find the issue string and description string from des string based on the starter_start_index above
            issue = des[starter_start_index:].strip()
            description = des[:starter_start_index - 1].strip()
            # insert the issue string to row (the row list)
            row.insert(3, issue)
            # replace description with the true description in the row (the row list)
            row[2] = description

        elif end_check(des)[0]:  # if the des string contain any end word
            # locate the start index of end in the des string
            end = end_check(des)[1]
            end_start_index = des.rfind(end+" ") # find the last occurrence index
            # find the issue string and description string from des string based on the end_start_index above
            issue = des[end_start_index + len(end):].strip()
            description = des[:end_start_index + len(end)].strip()
            # insert the issue string to row (the row list)
            row.insert(3, issue)
            # replace description with the true description in the row (the row list)
            row[2] = description

        else: # if no starter or end word is found, add a note to Issue column to ask the user
            # find issue type from description column manually
            note = "See description column for issue type"
            row.insert(3, note)
    return all_table

'''
Find total count number from 13F string
Input: el: the row string from split_list
Output: count number in integer
'''

def find_count(el):
    if "Total Count" in el:
        return [True, int(el[len(el)-6:].replace(',', ''))]
    return [False, 0]

'''
Format the saved Excel output:
1. remove index number column at column A
2. add filter for all columns
3. add count items and check against PDF total count
'''

def format_workbook(path, count, xlsx_count):
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]

    # delete number index column 1
    ws.delete_cols(1)

    # add count check box
    ws['G1'] = "Total Count Check:"
    ws['G2'] = "Per 13F PDF file:"
    ws['G3'] = "Per Converted Excel Here:"
    ws['G4'] = "Difference:"
    # calc total count difference between PDF and Excel files
    ws['H2'] = count
    ws['H3'] = xlsx_count
    ws['H4'] = "=H2-H3"

    # format count numbers and difference as comma style (Cells H2-H4)
    for i in range(1,5):
        ws['H'+str(i)].style = 'Comma [0]'

    # add filter
    ws.auto_filter.ref = f'A1:E{str(xlsx_count+1)}'

    # adjust column width
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # save workbook after all formatting work
    wb.save(path)

if __name__ == '__main__':
    # Obtain PDF file path
    input_path = obtain_pdf_file_path()

    # Read and organize PDF file into a list
    split_list = parsePDF(input_path)
    # Create a list to input all valid rows
    all_table = []
    # Create count to save total count result found from 13F file
    count = 0

    # Add valid table row as a list into all_table list
    for el in split_list:
        if find_count(el)[0]:
            count = find_count(el)[1] # update total count number from reading through all strings
        # Check what is in the el. If el contains CUSIP, it is a valid string, then save the string as a list
        # and append to all_table
        row_list = categorize_col_contents(el)
        if row_list != None:
            all_table.append(row_list)

    # Further split description and issue columns in all_table's lists
    final_table = split_table(all_table)

    # Add headings and save final_table as an Excel file
    df = pd.DataFrame(data=final_table)
    df = df.set_axis(['CUSIP', 'Option', "Description", "Issue", "Status"], axis=1)
    save_path = save_excel_file_path()
    df.to_excel(save_path)

    # Format the Excel file
    format_workbook(save_path, count, len(final_table))
